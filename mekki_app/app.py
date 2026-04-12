from flask import Flask, render_template, request, redirect, url_for, send_file, flash
import sqlite3
import os
from datetime import datetime
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment

app = Flask(__name__)
app.secret_key = "mekki-secret-key"
DB_PATH = "orders.db"

MEKKI_TYPES = [
    "ニッケルメッキ",
    "クロムメッキ",
    "亜鉛メッキ",
    "金メッキ",
    "銀メッキ",
    "銅メッキ",
    "無電解ニッケル",
    "硬質クロム",
    "スズメッキ",
    "三価クロメート",
    "クロメート",
    "黒黒メート",
    "その他",
]

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                order_no TEXT NOT NULL,
                customer TEXT NOT NULL,
                product TEXT NOT NULL,
                quantity INTEGER NOT NULL,
                mekki_type TEXT NOT NULL,
                due_date TEXT NOT NULL,
                note TEXT,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL
            )
        """)

@app.route("/")
def index():
    with get_db() as conn:
        orders = conn.execute(
            "SELECT * FROM orders ORDER BY created_at DESC"
        ).fetchall()
    return render_template("index.html", orders=orders)

@app.route("/new", methods=["GET", "POST"])
def new_order():
    if request.method == "POST":
        now = datetime.now()
        order_no = f"ORD-{now.strftime('%Y%m%d%H%M%S')}"
        with get_db() as conn:
            conn.execute("""
                INSERT INTO orders (order_no, customer, product, quantity, mekki_type, due_date, note, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                order_no,
                request.form["customer"],
                request.form["product"],
                request.form["quantity"],
                request.form["mekki_type"],
                request.form["due_date"],
                request.form.get("note", ""),
                now.strftime("%Y-%m-%d %H:%M:%S"),
            ))
        return redirect(url_for("index"))
    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
    return render_template("new.html", mekki_types=MEKKI_TYPES, customers=customers)

@app.route("/edit/<int:order_id>", methods=["GET", "POST"])
def edit_order(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return redirect(url_for("index"))
    if request.method == "POST":
        with get_db() as conn:
            conn.execute("""
                UPDATE orders SET customer=?, product=?, quantity=?,
                mekki_type=?, due_date=?, note=? WHERE id=?
            """, (
                request.form["customer"],
                request.form["product"],
                request.form["quantity"],
                request.form["mekki_type"],
                request.form["due_date"],
                request.form.get("note", ""),
                order_id,
            ))
        return redirect(url_for("detail", order_id=order_id))
    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
    return render_template("edit.html", order=order, mekki_types=MEKKI_TYPES, customers=customers)

@app.route("/detail/<int:order_id>")
def detail(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return redirect(url_for("index"))
    return render_template("detail.html", order=order)

@app.route("/print/<int:order_id>")
def print_order(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return redirect(url_for("index"))
    return render_template("print.html", order=order)

@app.route("/delete/<int:order_id>", methods=["POST"])
def delete(order_id):
    with get_db() as conn:
        conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
    return redirect(url_for("index"))

# ── 顧客マスタ ──────────────────────────────────────────────

@app.route("/customers", methods=["GET", "POST"])
def customers():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.endswith(".xlsx"):
            flash("xlsx ファイルを選択してください。", "error")
            return redirect(url_for("customers"))
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            imported = 0
            skipped = 0
            with get_db() as conn:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = str(row[0]).strip() if row[0] is not None else ""
                    if not name or name == "None":
                        continue
                    existing = conn.execute(
                        "SELECT id FROM customers WHERE name=?", (name,)
                    ).fetchone()
                    if existing:
                        skipped += 1
                    else:
                        conn.execute(
                            "INSERT INTO customers (name, created_at) VALUES (?, ?)",
                            (name, now)
                        )
                        imported += 1
            msg = f"{imported} 件を登録しました。"
            if skipped:
                msg += f"（{skipped} 件は既存のためスキップ）"
            flash(msg, "success")
        except Exception as e:
            flash(f"読み込みエラー: {e}", "error")
        return redirect(url_for("customers"))

    with get_db() as conn:
        customers_list = conn.execute(
            "SELECT * FROM customers ORDER BY name"
        ).fetchall()
    return render_template("customers.html", customers=customers_list)

@app.route("/customers/edit/<int:customer_id>", methods=["POST"])
def edit_customer(customer_id):
    name = request.form.get("name", "").strip()
    if name:
        try:
            with get_db() as conn:
                conn.execute("UPDATE customers SET name=? WHERE id=?", (name, customer_id))
            flash(f"顧客名を「{name}」に更新しました。", "success")
        except Exception:
            flash("同じ顧客名が既に登録されています。", "error")
    return redirect(url_for("customers"))

@app.route("/customers/delete/<int:customer_id>", methods=["POST"])
def delete_customer(customer_id):
    with get_db() as conn:
        conn.execute("DELETE FROM customers WHERE id=?", (customer_id,))
    return redirect(url_for("customers"))

@app.route("/customers/delete-all", methods=["POST"])
def delete_all_customers():
    with get_db() as conn:
        conn.execute("DELETE FROM customers")
    flash("顧客マスタを全件削除しました。", "success")
    return redirect(url_for("customers"))

# ── Excelエクスポート ────────────────────────────────────────

@app.route("/export/excel")
def export_excel():
    with get_db() as conn:
        orders = conn.execute(
            "SELECT * FROM orders ORDER BY created_at DESC"
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "受注データ"

    headers = ["伝票番号", "顧客名", "品名", "数量", "メッキ種類", "納期", "備考", "登録日時"]
    ws.append(headers)

    # ヘッダー行のスタイル
    header_fill = PatternFill(fill_type="solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # データ行
    for order in orders:
        ws.append([
            order["order_no"],
            order["customer"],
            order["product"],
            order["quantity"],
            order["mekki_type"],
            order["due_date"],
            order["note"] or "",
            order["created_at"],
        ])

    # 列幅の自動調整
    col_widths = [22, 20, 20, 8, 16, 14, 30, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

if __name__ == "__main__":
    init_db()
    app.run(debug=True)
