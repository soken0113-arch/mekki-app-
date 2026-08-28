from io import BytesIO

from flask import Blueprint, render_template, request, send_file
from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import func, select

from extensions import db
from models import DailyMenu, GuestOrder, MealType, Order, User, get_setting, get_setting_int
from timeutil import (
    add_months,
    fmt_date_full,
    fmt_date_short,
    parse_date,
    parse_month,
    period_label,
    period_of,
    period_range,
    today_jst,
)
from utils import admin_required, order_status

bp = Blueprint("reports", __name__, url_prefix="/reports")


# ── 日次（仕出し屋への発注用） ──────────────────────────────

def _daily_summary(serve_date):
    """その日の区分別 食数（社員・来客・合計）。提供する区分は 0 食でも行を残す。"""
    menus = (
        db.session.execute(
            select(DailyMenu)
            .join(MealType)
            .where(DailyMenu.serve_date == serve_date)
            .order_by(MealType.sort_order)
        )
        .scalars()
        .all()
    )

    staff_counts = dict(
        db.session.execute(
            select(Order.meal_type_id, func.count(Order.id))
            .where(Order.serve_date == serve_date)
            .group_by(Order.meal_type_id)
        ).all()
    )
    guest_counts = dict(
        db.session.execute(
            select(GuestOrder.meal_type_id, GuestOrder.count).where(
                GuestOrder.serve_date == serve_date
            )
        ).all()
    )

    rows = []
    for menu in menus:
        staff = staff_counts.get(menu.meal_type_id, 0)
        guest = guest_counts.get(menu.meal_type_id, 0)
        rows.append(
            {
                "meal_type_id": menu.meal_type_id,
                "name": menu.meal_type.name,
                "dish": menu.dish_name,
                "staff": staff,
                "guest": guest,
                "total": staff + guest,
            }
        )
    return rows


@bp.route("/daily")
@admin_required
def daily():
    serve_date = parse_date(request.args.get("date", ""))
    summary = _daily_summary(serve_date)

    orders = (
        Order.query.join(User, User.id == Order.user_id)
        .filter(Order.serve_date == serve_date)
        .order_by(User.sort_order, User.name)
        .all()
    )
    ordered_ids = {o.user_id for o in orders}
    not_ordered = [
        u
        for u in User.query.filter_by(is_active=True).order_by(User.sort_order, User.name).all()
        if u.id not in ordered_ids
    ]

    return render_template(
        "report_daily.html",
        serve_date=serve_date,
        summary=summary,
        orders=orders,
        not_ordered=not_ordered,
        total_staff=sum(r["staff"] for r in summary),
        total_guest=sum(r["guest"] for r in summary),
        total_all=sum(r["total"] for r in summary),
        users=User.query.filter_by(is_active=True).order_by(User.sort_order, User.name).all(),
        shop_name=get_setting("shop_name"),
        status=order_status(serve_date),
        today=today_jst(),
    )


@bp.route("/daily/export.xlsx")
@admin_required
def daily_export():
    serve_date = parse_date(request.args.get("date", ""))
    summary = _daily_summary(serve_date)
    shop_name = get_setting("shop_name")

    wb = Workbook()
    ws = wb.active
    ws.title = "発注書"
    ws.append([f"昼食発注書　{fmt_date_full(serve_date)}"])
    ws["A1"].font = Font(bold=True, size=14)
    if shop_name:
        ws.append([f"発注先: {shop_name}"])
    ws.append([])
    ws.append(["区分", "献立", "社員", "来客", "合計"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for row in summary:
        ws.append([row["name"], row["dish"], row["staff"], row["guest"], row["total"]])
    ws.append([
        "合計", "",
        sum(r["staff"] for r in summary),
        sum(r["guest"] for r in summary),
        sum(r["total"] for r in summary),
    ])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    ws2 = wb.create_sheet("社員別内訳")
    ws2.append(["社員番号", "部署", "氏名", "区分", "献立"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    orders = (
        Order.query.join(User, User.id == Order.user_id)
        .filter(Order.serve_date == serve_date)
        .order_by(User.sort_order, User.name)
        .all()
    )
    dishes = {
        m.meal_type_id: m.dish_name
        for m in DailyMenu.query.filter_by(serve_date=serve_date).all()
    }
    for order in orders:
        ws2.append(
            [
                order.user.employee_no,
                order.user.department,
                order.user.name,
                order.meal_type.name,
                dishes.get(order.meal_type_id, ""),
            ]
        )

    _autosize(ws)
    _autosize(ws2)
    return _xlsx_response(wb, f"lunch_order_{serve_date:%Y%m%d}.xlsx")


# ── 月次（個人別の食数） ────────────────────────────────────

def _period_from_request():
    start_day = get_setting_int("period_start_day", 21)
    month = parse_month(
        request.args.get("month", ""), default=period_of(today_jst(), start_day)
    )
    start, end = period_range(month, start_day)
    return month, start, end, start_day


def _monthly_rows(start, end):
    """社員ごとの食数と区分別内訳。"""
    types = MealType.query.order_by(MealType.sort_order).all()
    counts = db.session.execute(
        select(Order.user_id, Order.meal_type_id, func.count(Order.id))
        .where(Order.serve_date >= start, Order.serve_date < end)
        .group_by(Order.user_id, Order.meal_type_id)
    ).all()

    per_user: dict = {}
    for user_id, meal_type_id, count in counts:
        per_user.setdefault(user_id, {})[meal_type_id] = count

    rows = []
    for user in User.query.order_by(User.sort_order, User.name).all():
        breakdown = per_user.get(user.id, {})
        total = sum(breakdown.values())
        if not user.is_active and total == 0:
            continue  # 退職者などは、その期間に注文があるときだけ出す
        rows.append({"user": user, "breakdown": breakdown, "total": total})
    return types, rows


def _guest_totals(start, end):
    return dict(
        db.session.execute(
            select(GuestOrder.meal_type_id, func.sum(GuestOrder.count))
            .where(GuestOrder.serve_date >= start, GuestOrder.serve_date < end)
            .group_by(GuestOrder.meal_type_id)
        ).all()
    )


@bp.route("/monthly")
@admin_required
def monthly():
    month, start, end, start_day = _period_from_request()
    types, rows = _monthly_rows(start, end)
    guests = _guest_totals(start, end)

    by_date = db.session.execute(
        select(Order.serve_date, func.count(Order.id))
        .where(Order.serve_date >= start, Order.serve_date < end)
        .group_by(Order.serve_date)
        .order_by(Order.serve_date)
    ).all()
    guest_by_date = dict(
        db.session.execute(
            select(GuestOrder.serve_date, func.sum(GuestOrder.count))
            .where(GuestOrder.serve_date >= start, GuestOrder.serve_date < end)
            .group_by(GuestOrder.serve_date)
        ).all()
    )

    return render_template(
        "report_monthly.html",
        month=month,
        period_text=period_label(month, start_day),
        types=types,
        rows=rows,
        guests=guests,
        guest_total=sum(guests.values()),
        total_count=sum(r["total"] for r in rows),
        type_totals={t.id: sum(r["breakdown"].get(t.id, 0) for r in rows) for t in types},
        by_date=[
            {"date": d, "staff": c, "guest": guest_by_date.get(d, 0), "total": c + guest_by_date.get(d, 0)}
            for d, c in by_date
        ],
        prev_month=add_months(month, -1),
        next_month=add_months(month, 1),
    )


@bp.route("/monthly/export.xlsx")
@admin_required
def monthly_export():
    month, start, end, start_day = _period_from_request()
    types, rows = _monthly_rows(start, end)
    guests = _guest_totals(start, end)

    wb = Workbook()
    ws = wb.active
    ws.title = "社員別集計"
    ws.append([f"昼食 個人別集計　{period_label(month, start_day)}"])
    ws["A1"].font = Font(bold=True, size=14)
    ws.append([])
    ws.append(["社員番号", "部署", "氏名"] + [t.name for t in types] + ["食数合計"])
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)
    for row in rows:
        user = row["user"]
        ws.append(
            [user.employee_no, user.department, user.name]
            + [row["breakdown"].get(t.id, 0) for t in types]
            + [row["total"]]
        )
    ws.append(
        ["", "", "来客用"]
        + [guests.get(t.id, 0) for t in types]
        + [sum(guests.values())]
    )
    ws.append(
        ["", "", "合計"]
        + [sum(r["breakdown"].get(t.id, 0) for r in rows) + guests.get(t.id, 0) for t in types]
        + [sum(r["total"] for r in rows) + sum(guests.values())]
    )
    for cell in ws[ws.max_row]:
        cell.font = Font(bold=True)

    ws2 = wb.create_sheet("日別明細")
    ws2.append(["日付", "社員番号", "氏名", "区分"])
    for cell in ws2[1]:
        cell.font = Font(bold=True)
    detail = (
        Order.query.join(User, User.id == Order.user_id)
        .filter(Order.serve_date >= start, Order.serve_date < end)
        .order_by(Order.serve_date, User.sort_order, User.name)
        .all()
    )
    for order in detail:
        ws2.append(
            [
                order.serve_date.strftime("%Y-%m-%d"),
                order.user.employee_no,
                order.user.name,
                order.meal_type.name,
            ]
        )

    _autosize(ws)
    _autosize(ws2)
    return _xlsx_response(wb, f"lunch_monthly_{month:%Y%m}.xlsx")


# ── 共通 ────────────────────────────────────────────────────

def _autosize(ws) -> None:
    for column in ws.columns:
        width = max((len(str(c.value)) for c in column if c.value is not None), default=0)
        ws.column_dimensions[column[0].column_letter].width = min(max(width + 4, 9), 40)


def _xlsx_response(wb, filename):
    stream = BytesIO()
    wb.save(stream)
    stream.seek(0)
    return send_file(
        stream,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
