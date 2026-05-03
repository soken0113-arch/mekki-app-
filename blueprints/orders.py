from flask import Blueprint, render_template, request, redirect, url_for, send_file, flash
from datetime import datetime
from io import BytesIO
import uuid
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from db import get_db
from utils import (login_required, MEKKI_TYPES, MEKKI_LINES,
                   build_thickness, parse_thickness, get_jp_holidays, _build_order_list)

bp = Blueprint("orders", __name__)


@bp.route("/")
@login_required
def index():
    with get_db() as conn:
        rows = conn.execute(
            "SELECT * FROM orders WHERE (subcontractor_id IS NULL) ORDER BY created_at DESC"
        ).fetchall()
        shipped_rows = conn.execute("SELECT order_id FROM shipments").fetchall()
    shipped_ids = {r["order_id"] for r in shipped_rows}
    orders = _build_order_list(rows, shipped_ids, get_jp_holidays(), datetime.now().date())
    alert_orders = [o for o in orders if o['alert']]
    return render_template("index.html", orders=orders, alert_orders=alert_orders)


@bp.route("/sub_orders")
@login_required
def sub_orders_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT o.*, s.name as subcontractor_name
            FROM orders o
            LEFT JOIN subcontractors s ON o.subcontractor_id = s.id
            WHERE o.subcontractor_id IS NOT NULL
            ORDER BY o.created_at DESC
        """).fetchall()
        shipped_rows = conn.execute("SELECT order_id FROM shipments").fetchall()
    shipped_ids = {r["order_id"] for r in shipped_rows}
    orders = _build_order_list(rows, shipped_ids, get_jp_holidays(), datetime.now().date())
    return render_template("sub_orders.html", orders=orders)


@bp.route("/new", methods=["GET", "POST"])
@login_required
def new_order():
    if request.method == "POST":
        now = datetime.now()
        order_no = f"ORD-{now.strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}"
        with get_db() as conn:
            cur = conn.execute("""
                INSERT INTO orders (order_no, customer, product, part_no, material, quantity,
                    mekki_type, mekki_thickness, thickness_data, due_date,
                    unit_price, mekki_line, process_note, shipping_method, note, assigned_to, created_at,
                    subcontractor_id)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                RETURNING id
            """, (
                order_no,
                request.form["customer"],
                request.form["product"],
                request.form.get("part_no", ""),
                request.form.get("material", ""),
                request.form["quantity"],
                request.form["mekki_type"],
                build_thickness(request.form.get("thickness_from", ""), request.form.get("thickness_to", "")),
                request.form.get("thickness_data", "不要"),
                request.form["due_date"],
                request.form.get("unit_price", ""),
                request.form.get("mekki_line", ""),
                request.form.get("process_note", ""),
                request.form.get("shipping_method", ""),
                request.form.get("note", ""),
                request.form.get("assigned_to", ""),
                now.strftime("%Y-%m-%d %H:%M:%S"),
                request.form.get("subcontractor_id") or None,
            ))
            new_id = cur.fetchone()["id"]
        return redirect(url_for("orders.detail", order_id=new_id))
    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
        products = conn.execute("SELECT name, part_no, unit_price, note FROM products ORDER BY name").fetchall()
        subcontractors = conn.execute("SELECT id, name FROM subcontractors ORDER BY name").fetchall()
    return render_template("new.html", mekki_types=MEKKI_TYPES, mekki_lines=MEKKI_LINES,
                           customers=customers, products=products, subcontractors=subcontractors)


@bp.route("/edit/<int:order_id>", methods=["GET", "POST"])
@login_required
def edit_order(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return redirect(url_for("orders.index"))
    if request.method == "POST":
        with get_db() as conn:
            conn.execute("""
                UPDATE orders SET customer=?, product=?, part_no=?, material=?, quantity=?,
                mekki_type=?, mekki_thickness=?, thickness_data=?, due_date=?,
                unit_price=?, mekki_line=?, process_note=?, shipping_method=?, note=?, assigned_to=?,
                subcontractor_id=?
                WHERE id=?
            """, (
                request.form["customer"],
                request.form["product"],
                request.form.get("part_no", ""),
                request.form.get("material", ""),
                request.form["quantity"],
                request.form["mekki_type"],
                build_thickness(request.form.get("thickness_from", ""), request.form.get("thickness_to", "")),
                request.form.get("thickness_data", "不要"),
                request.form["due_date"],
                request.form.get("unit_price", ""),
                request.form.get("mekki_line", ""),
                request.form.get("process_note", ""),
                request.form.get("shipping_method", ""),
                request.form.get("note", ""),
                request.form.get("assigned_to", ""),
                request.form.get("subcontractor_id") or None,
                order_id,
            ))
        return redirect(url_for("orders.detail", order_id=order_id))
    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
        products = conn.execute("SELECT name, part_no, unit_price, note FROM products ORDER BY name").fetchall()
        subcontractors = conn.execute("SELECT id, name FROM subcontractors ORDER BY name").fetchall()
    t_from, t_to = parse_thickness(order["mekki_thickness"])
    return render_template("edit.html", order=order, mekki_types=MEKKI_TYPES, mekki_lines=MEKKI_LINES,
                           customers=customers, products=products, subcontractors=subcontractors,
                           thickness_from=t_from, thickness_to=t_to)


@bp.route("/detail/<int:order_id>")
@login_required
def detail(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            return redirect(url_for("orders.index"))
        shipment = conn.execute("SELECT * FROM shipments WHERE order_id=?", (order_id,)).fetchone()
        subcontractor_name = None
        try:
            sub_id = order["subcontractor_id"]
            if sub_id:
                sub = conn.execute("SELECT name FROM subcontractors WHERE id=?", (sub_id,)).fetchone()
                subcontractor_name = sub["name"] if sub else None
        except Exception:
            pass
    return render_template("detail.html", order=order, shipment=shipment, subcontractor_name=subcontractor_name)


@bp.route("/print/<int:order_id>")
@login_required
def print_order(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return redirect(url_for("orders.index"))
    return render_template("print.html", order=order)


@bp.route("/delete/<int:order_id>", methods=["POST"])
@login_required
def delete(order_id):
    with get_db() as conn:
        conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
    referrer = request.referrer or ""
    if "sub_orders" in referrer:
        return redirect(url_for("orders.sub_orders_list"))
    return redirect(url_for("orders.index"))


@bp.route("/new_multi", methods=["GET", "POST"])
@login_required
def new_order_multi():
    if request.method == "POST":
        now = datetime.now()
        order_no = f"ORD-{now.strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}-M"
        with get_db() as conn:
            cur = conn.execute("""
                INSERT INTO orders (
                    order_no, customer, product, part_no, material, quantity,
                    mekki_type, mekki_thickness, thickness_data, due_date,
                    unit_price, mekki_line, process_note, shipping_method, note, assigned_to, created_at,
                    subcontractor_id
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                RETURNING id
            """, (
                order_no,
                request.form["customer"],
                "複数品目",
                "",
                request.form.get("material", ""),
                0,
                request.form["mekki_type"],
                build_thickness(request.form.get("thickness_from", ""), request.form.get("thickness_to", "")),
                request.form.get("thickness_data", "不要"),
                request.form["due_date"],
                "",
                request.form.get("mekki_line", ""),
                request.form.get("process_note", ""),
                request.form.get("shipping_method", ""),
                request.form.get("note", ""),
                request.form.get("assigned_to", ""),
                now.strftime("%Y-%m-%d %H:%M:%S"),
                request.form.get("subcontractor_id") or None,
            ))
            order_id = cur.fetchone()["id"]

            products_list = request.form.getlist("product[]")
            part_nos = request.form.getlist("part_no[]")
            materials = request.form.getlist("material[]")
            quantities = request.form.getlist("quantity[]")
            unit_prices = request.form.getlist("unit_price[]")
            notes = request.form.getlist("note[]")
            for i, product in enumerate(products_list):
                if not product.strip():
                    continue
                conn.execute("""
                    INSERT INTO order_items (
                        order_id, part_no, product, material, quantity, unit_price, note
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    order_id,
                    part_nos[i] if i < len(part_nos) else "",
                    product,
                    materials[i] if i < len(materials) else "",
                    int(quantities[i]) if i < len(quantities) and quantities[i] else 0,
                    unit_prices[i] if i < len(unit_prices) else "",
                    notes[i] if i < len(notes) else "",
                ))
        return redirect(url_for("orders.detail_multi", order_id=order_id))
    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
        products = conn.execute("SELECT name, part_no, unit_price, note FROM products ORDER BY name").fetchall()
        subcontractors = conn.execute("SELECT id, name FROM subcontractors ORDER BY name").fetchall()
    return render_template("new_multi.html", mekki_types=MEKKI_TYPES, mekki_lines=MEKKI_LINES,
                           customers=customers, products=products, subcontractors=subcontractors)


@bp.route("/edit_multi/<int:order_id>", methods=["GET", "POST"])
@login_required
def edit_order_multi(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        items = conn.execute(
            "SELECT * FROM order_items WHERE order_id=? ORDER BY id", (order_id,)
        ).fetchall()
    if not order:
        return redirect(url_for("orders.index"))
    if request.method == "POST":
        with get_db() as conn:
            conn.execute("""
                UPDATE orders SET
                    customer=?, mekki_type=?, mekki_thickness=?, thickness_data=?,
                    material=?, due_date=?, mekki_line=?, process_note=?,
                    shipping_method=?, note=?, assigned_to=?,
                    subcontractor_id=?
                WHERE id=?
            """, (
                request.form["customer"],
                request.form["mekki_type"],
                build_thickness(request.form.get("thickness_from", ""), request.form.get("thickness_to", "")),
                request.form.get("thickness_data", "不要"),
                request.form.get("material", ""),
                request.form["due_date"],
                request.form.get("mekki_line", ""),
                request.form.get("process_note", ""),
                request.form.get("shipping_method", ""),
                request.form.get("note", ""),
                request.form.get("assigned_to", ""),
                request.form.get("subcontractor_id") or None,
                order_id,
            ))
            conn.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
            products_list = request.form.getlist("product[]")
            part_nos = request.form.getlist("part_no[]")
            materials = request.form.getlist("material[]")
            quantities = request.form.getlist("quantity[]")
            unit_prices = request.form.getlist("unit_price[]")
            notes = request.form.getlist("note[]")
            for i, product in enumerate(products_list):
                if not product.strip():
                    continue
                conn.execute("""
                    INSERT INTO order_items (
                        order_id, part_no, product, material, quantity, unit_price, note
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    order_id,
                    part_nos[i] if i < len(part_nos) else "",
                    product,
                    materials[i] if i < len(materials) else "",
                    int(quantities[i]) if i < len(quantities) and quantities[i] else 0,
                    unit_prices[i] if i < len(unit_prices) else "",
                    notes[i] if i < len(notes) else "",
                ))
        return redirect(url_for("orders.detail_multi", order_id=order_id))
    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
        products = conn.execute("SELECT name, part_no, unit_price, note FROM products ORDER BY name").fetchall()
        subcontractors = conn.execute("SELECT id, name FROM subcontractors ORDER BY name").fetchall()
    t_from, t_to = parse_thickness(order["mekki_thickness"])
    return render_template("edit_multi.html", order=order, items=items,
                           mekki_types=MEKKI_TYPES, mekki_lines=MEKKI_LINES,
                           customers=customers, products=products, subcontractors=subcontractors,
                           thickness_from=t_from, thickness_to=t_to)


@bp.route("/detail_multi/<int:order_id>")
@login_required
def detail_multi(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            return redirect(url_for("orders.index"))
        items = conn.execute(
            "SELECT * FROM order_items WHERE order_id=? ORDER BY id", (order_id,)
        ).fetchall()
        shipment = conn.execute("SELECT * FROM shipments WHERE order_id=?", (order_id,)).fetchone()
        subcontractor_name = None
        try:
            sub_id = order["subcontractor_id"]
            if sub_id:
                sub = conn.execute("SELECT name FROM subcontractors WHERE id=?", (sub_id,)).fetchone()
                subcontractor_name = sub["name"] if sub else None
        except Exception:
            pass
    return render_template("detail_multi.html", order=order, items=items,
                           shipment=shipment, subcontractor_name=subcontractor_name)


@bp.route("/print_multi/<int:order_id>")
@login_required
def print_order_multi(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        items = conn.execute(
            "SELECT * FROM order_items WHERE order_id=? ORDER BY id", (order_id,)
        ).fetchall()
    if not order:
        return redirect(url_for("orders.index"))
    return render_template("print_multi.html", order=order, items=items)


@bp.route("/new_gaichuu", methods=["GET", "POST"])
@login_required
def new_order_gaichuu():
    if request.method == "POST":
        now = datetime.now()
        order_no = f"ORD-{now.strftime('%Y%m%d%H%M%S')}-{uuid.uuid4().hex[:6].upper()}-G"
        with get_db() as conn:
            cur = conn.execute("""
                INSERT INTO orders (
                    order_no, customer, product, part_no, material, quantity,
                    mekki_type, mekki_thickness, thickness_data, due_date,
                    unit_price, mekki_line, process_note, shipping_method, note, assigned_to, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                RETURNING id
            """, (
                order_no,
                request.form["customer"],
                "複数品目",
                "",
                request.form.get("material", ""),
                0,
                request.form.get("mekki_type", ""),
                build_thickness(request.form.get("thickness_from", ""), request.form.get("thickness_to", "")),
                request.form.get("thickness_data", "不要"),
                request.form["due_date"],
                "",
                "外注",
                request.form.get("process_note", ""),
                request.form.get("shipping_method", ""),
                request.form.get("note", ""),
                request.form.get("assigned_to", ""),
                now.strftime("%Y-%m-%d %H:%M:%S"),
            ))
            order_id = cur.fetchone()["id"]

            products_list = request.form.getlist("product[]")
            part_nos = request.form.getlist("part_no[]")
            materials = request.form.getlist("material[]")
            quantities = request.form.getlist("quantity[]")
            unit_prices = request.form.getlist("unit_price[]")
            notes = request.form.getlist("note[]")
            for i, product in enumerate(products_list):
                if not product.strip():
                    continue
                conn.execute("""
                    INSERT INTO order_items (
                        order_id, part_no, product, material, quantity, unit_price, note
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    order_id,
                    part_nos[i] if i < len(part_nos) else "",
                    product,
                    materials[i] if i < len(materials) else "",
                    int(quantities[i]) if i < len(quantities) and quantities[i] else 0,
                    unit_prices[i] if i < len(unit_prices) else "",
                    notes[i] if i < len(notes) else "",
                ))
        return redirect(url_for("orders.index"))
    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
        products = conn.execute("SELECT name, part_no, unit_price, note FROM products ORDER BY name").fetchall()
    return render_template("new_gaichuu.html", mekki_types=MEKKI_TYPES,
                           customers=customers, products=products)


@bp.route("/export/excel")
@login_required
def export_excel():
    with get_db() as conn:
        orders = conn.execute("SELECT * FROM orders ORDER BY created_at DESC").fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "受注データ"

    headers = ["伝票番号", "顧客名", "品名", "品番", "材質", "数量", "めっき種類", "めっき膜厚",
               "膜厚データ", "納期", "単価", "めっきライン", "工程", "出荷方法", "備考", "登録日時"]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for order in orders:
        ws.append([
            order["order_no"], order["customer"], order["product"],
            order["part_no"] or "", order["material"] or "", order["quantity"],
            order["mekki_type"], order["mekki_thickness"] or "",
            order["thickness_data"] or "", order["due_date"],
            order["unit_price"] or "", order["mekki_line"] or "",
            order["process_note"] or "", order["shipping_method"] or "",
            order["note"] or "", order["created_at"],
        ])

    col_widths = [22, 20, 20, 16, 16, 8, 16, 14, 10, 14, 10, 16, 24, 12, 30, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
