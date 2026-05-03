from flask import Blueprint, render_template, request, redirect, url_for, flash, send_file
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from db import get_db
from utils import login_required, _register_shipment

bp = Blueprint("shipments", __name__)


@bp.route("/shipments/complete/<int:order_id>", methods=["POST"])
@login_required
def complete_shipment(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            return redirect(url_for("orders.index"))
        registered = _register_shipment(conn, order_id, order)
    flash("出荷完了として登録しました。" if registered else "この受注はすでに出荷済みです。",
          "success" if registered else "info")
    return redirect(url_for("orders.index"))


@bp.route("/ship/<int:order_id>", methods=["POST"])
@login_required
def ship_order(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
        if not order:
            return redirect(url_for("orders.index"))
        registered = _register_shipment(conn, order_id, order)
    flash("出荷完了として登録しました。" if registered else "この受注はすでに出荷済みです。",
          "success" if registered else "info")
    referrer = request.referrer or ""
    if "sub_orders" in referrer:
        return redirect(url_for("orders.sub_orders_list"))
    if order["product"] == "複数品目":
        return redirect(url_for("orders.detail_multi", order_id=order_id))
    return redirect(url_for("orders.detail", order_id=order_id))


@bp.route("/shipments")
@login_required
def shipments_list():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT s.id, s.order_id, s.shipped_at, s.shipped_by, s.note,
                   o.order_no, o.customer
            FROM shipments s
            JOIN orders o ON s.order_id = o.id
            ORDER BY s.shipped_at DESC
        """).fetchall()
    today = datetime.now().date()
    cutoff = today - timedelta(days=90)
    shipments = []
    for row in rows:
        s = dict(row)
        try:
            shipped_date = datetime.strptime(s["shipped_at"], "%Y-%m-%d %H:%M:%S").date()
            s["is_archived"] = shipped_date < cutoff
        except Exception:
            s["is_archived"] = False
        shipments.append(s)
    return render_template("shipments.html", shipments=shipments)


@bp.route("/shipments/delete/<int:shipment_id>", methods=["POST"])
@login_required
def delete_shipment(shipment_id):
    with get_db() as conn:
        shipment = conn.execute("SELECT order_id FROM shipments WHERE id=?", (shipment_id,)).fetchone()
        if shipment:
            order_id = shipment["order_id"]
            conn.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
            conn.execute("DELETE FROM shipments WHERE id=?", (shipment_id,))
            conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
    flash("出荷レコードと受注データを削除しました。", "success")
    return redirect(url_for("shipments.shipments_list"))


@bp.route("/shipments/export/excel")
@login_required
def export_shipments_excel():
    with get_db() as conn:
        rows = conn.execute("""
            SELECT s.shipped_at, s.shipped_by, s.note,
                   o.order_no, o.customer
            FROM shipments s
            JOIN orders o ON s.order_id = o.id
            ORDER BY s.shipped_at DESC
        """).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "出荷データ"

    headers = ["伝票番号", "得意先名", "出荷日時", "担当者", "備考"]
    ws.append(headers)

    header_fill = PatternFill(fill_type="solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    for row in rows:
        ws.append([
            row["order_no"], row["customer"], row["shipped_at"],
            row["shipped_by"] or "", row["note"] or "",
        ])

    col_widths = [22, 20, 20, 16, 30]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)
    filename = f"shipments_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buf, as_attachment=True, download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
