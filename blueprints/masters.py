from flask import Blueprint, render_template, request, redirect, url_for, flash, jsonify
from datetime import datetime
import openpyxl
import psycopg2
from db import get_db
from utils import login_required

bp = Blueprint("masters", __name__)


# ── 顧客マスタ ──────────────────────────────────────────────

@bp.route("/customers", methods=["GET", "POST"])
@login_required
def customers():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.endswith(".xlsx"):
            flash("xlsx ファイルを選択してください。", "error")
            return redirect(url_for("masters.customers"))
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            imported = 0
            skipped = 0
            with get_db() as conn:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = str(row[0]).strip() if row[0] is not None else ""
                    if not name or name == "None":
                        continue
                    existing = conn.execute(
                        "SELECT id FROM customers WHERE name=?", (name,)
                    ).fetchone()
                    if existing:
                        skipped += 1
                    else:
                        conn.execute(
                            "INSERT INTO customers (name, created_at) VALUES (?, ?)", (name, now)
                        )
                        imported += 1
            msg = f"{imported} 件を登録しました。"
            if skipped:
                msg += f"（{skipped} 件は既存のためスキップ）"
            flash(msg, "success")
        except Exception as e:
            flash(f"読み込みエラー: {e}", "error")
        return redirect(url_for("masters.customers"))

    with get_db() as conn:
        customers_list = conn.execute("SELECT * FROM customers ORDER BY name").fetchall()
    return render_template("customers.html", customers=customers_list)


@bp.route("/customers/add", methods=["POST"])
@login_required
def add_customer():
    name = request.form.get("name", "").strip()
    if not name:
        return jsonify({"success": False, "error": "顧客名が空です"}), 400
    try:
        with get_db() as conn:
            row = conn.execute(
                "INSERT INTO customers (name, created_at) VALUES (?, ?) RETURNING id",
                (name, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ).fetchone()
        return jsonify({"success": True, "id": row["id"], "name": name})
    except Exception:
        return jsonify({"success": False, "error": "この顧客名はすでに登録されています"}), 409


@bp.route("/customers/edit/<int:customer_id>", methods=["POST"])
@login_required
def edit_customer(customer_id):
    name = request.form.get("name", "").strip()
    if name:
        try:
            with get_db() as conn:
                conn.execute("UPDATE customers SET name=? WHERE id=?", (name, customer_id))
            flash(f"顧客名を「{name}」に更新しました。", "success")
        except Exception:
            flash("同じ顧客名が既に登録されています。", "error")
    return redirect(url_for("masters.customers"))


@bp.route("/customers/delete/<int:customer_id>", methods=["POST"])
@login_required
def delete_customer(customer_id):
    with get_db() as conn:
        conn.execute("DELETE FROM customers WHERE id=?", (customer_id,))
    return redirect(url_for("masters.customers"))


@bp.route("/customers/delete-all", methods=["POST"])
@login_required
def delete_all_customers():
    with get_db() as conn:
        conn.execute("DELETE FROM customers")
    flash("顧客マスタを全件削除しました。", "success")
    return redirect(url_for("masters.customers"))


# ── 品名マスタ ──────────────────────────────────────────────

@bp.route("/products", methods=["GET", "POST"])
@login_required
def products():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.endswith(".xlsx"):
            flash("xlsx ファイルを選択してください。", "error")
            return redirect(url_for("masters.products"))
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            imported = 0
            skipped = 0
            with get_db() as conn:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = str(row[0]).strip() if row[0] is not None else ""
                    if not name or name == "None":
                        continue
                    existing = conn.execute(
                        "SELECT id FROM products WHERE name=?", (name,)
                    ).fetchone()
                    if existing:
                        skipped += 1
                    else:
                        conn.execute(
                            "INSERT INTO products (name, created_at) VALUES (?, ?)", (name, now)
                        )
                        imported += 1
            msg = f"{imported} 件を登録しました。"
            if skipped:
                msg += f"（{skipped} 件は既存のためスキップ）"
            flash(msg, "success")
        except Exception as e:
            flash(f"読み込みエラー: {e}", "error")
        return redirect(url_for("masters.products"))

    with get_db() as conn:
        products_list = conn.execute("SELECT * FROM products ORDER BY name").fetchall()
    return render_template("products.html", products=products_list)


@bp.route("/products/add", methods=["POST"])
@login_required
def add_product():
    name = request.form.get("name", "").strip()
    if not name:
        return jsonify({"success": False, "error": "品名が空です"}), 400
    part_no = request.form.get("part_no", "").strip()
    unit_price = request.form.get("unit_price", "").strip()
    note = request.form.get("note", "").strip()
    try:
        with get_db() as conn:
            row = conn.execute(
                "INSERT INTO products (name, part_no, unit_price, note, created_at) VALUES (?, ?, ?, ?, ?) RETURNING id",
                (name, part_no, unit_price, note, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ).fetchone()
        return jsonify({"success": True, "id": row["id"], "name": name, "part_no": part_no})
    except psycopg2.IntegrityError:
        return jsonify({"success": False, "error": "この品名はすでに登録されています"}), 409
    except Exception as e:
        return jsonify({"success": False, "error": f"DB エラー: {e}"}), 500


@bp.route("/products/edit/<int:product_id>", methods=["POST"])
@login_required
def edit_product(product_id):
    name = request.form.get("name", "").strip()
    if name:
        try:
            with get_db() as conn:
                conn.execute("UPDATE products SET name=?, part_no=?, unit_price=?, note=? WHERE id=?", (
                    name,
                    request.form.get("part_no", "").strip(),
                    request.form.get("unit_price", "").strip(),
                    request.form.get("note", "").strip(),
                    product_id,
                ))
            flash(f"品名を「{name}」に更新しました。", "success")
        except Exception:
            flash("同じ品名が既に登録されています。", "error")
    return redirect(url_for("masters.products"))


@bp.route("/products/delete/<int:product_id>", methods=["POST"])
@login_required
def delete_product(product_id):
    with get_db() as conn:
        conn.execute("DELETE FROM products WHERE id=?", (product_id,))
    return redirect(url_for("masters.products"))


@bp.route("/products/delete-all", methods=["POST"])
@login_required
def delete_all_products():
    with get_db() as conn:
        conn.execute("DELETE FROM products")
    flash("品名マスタを全件削除しました。", "success")
    return redirect(url_for("masters.products"))


# ── 外注先マスタ ────────────────────────────────────────────

@bp.route("/subcontractors", methods=["GET"])
@login_required
def subcontractors():
    with get_db() as conn:
        subs = conn.execute("SELECT * FROM subcontractors ORDER BY name").fetchall()
    return render_template("subcontractors.html", subcontractors=subs)


@bp.route("/subcontractors/add", methods=["POST"])
@login_required
def add_subcontractor():
    name = request.form.get("name", "").strip()
    if not name:
        return jsonify({"success": False, "error": "外注先名が空です"}), 400
    try:
        with get_db() as conn:
            row = conn.execute(
                "INSERT INTO subcontractors (name, created_at) VALUES (?, ?) RETURNING id",
                (name, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            ).fetchone()
        return jsonify({"success": True, "id": row["id"], "name": name})
    except Exception:
        return jsonify({"success": False, "error": "この外注先名はすでに登録されています"}), 409


@bp.route("/subcontractors/edit/<int:sub_id>", methods=["POST"])
@login_required
def edit_subcontractor(sub_id):
    name = request.form.get("name", "").strip()
    if name:
        with get_db() as conn:
            conn.execute("UPDATE subcontractors SET name=? WHERE id=?", (name, sub_id))
    return redirect(url_for("masters.subcontractors"))


@bp.route("/subcontractors/delete/<int:sub_id>", methods=["POST"])
@login_required
def delete_subcontractor(sub_id):
    with get_db() as conn:
        conn.execute("DELETE FROM subcontractors WHERE id=?", (sub_id,))
    return redirect(url_for("masters.subcontractors"))
