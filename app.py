from flask import Flask, render_template, request, redirect, url_for, send_file, flash, session
from functools import wraps
import os
from datetime import datetime, timedelta
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment
from werkzeug.security import generate_password_hash, check_password_hash
import psycopg2
import psycopg2.extras

app = Flask(__name__)
app.secret_key = os.environ.get("SECRET_KEY", "mekki-secret-key")
app.permanent_session_lifetime = timedelta(minutes=30)
DATABASE_URL = os.environ.get("DATABASE_URL")
SESSION_TIMEOUT = 30  # 分

MEKKI_TYPES = [
    "ニッケルメッキ",
    "クロムメッキ",
    "亜鉛メッキ",
    "金メッキ",
    "銀メッキ",
    "銅メッキ",
    "無電解ニッケル",
    "硬質クロム",
    "スズメッキ",
    "三価クロメート",
    "クロメート",
    "黒黒メート",
    "その他",
]

MEKKI_LINES = ["銅", "２号機", "無Ⅰ", "無Ⅱ", "特殊", "アルマイト", "化成処理", "亜鉛", "Ni自動バレル", "外注"]

def build_thickness(from_val, to_val):
    """下限・上限を '下限〜上限μm' 形式の文字列にまとめる"""
    f = (from_val or "").strip()
    t = (to_val or "").strip()
    if f and t:
        return f"{f}〜{t}μm"
    elif f:
        return f"{f}μm"
    return ""

def parse_thickness(value):
    """'X〜Yμm' を (X, Y) に分解して返す。単独値は (X, '') を返す"""
    if not value:
        return "", ""
    v = value.replace("μm", "").strip()
    if "〜" in v:
        parts = v.split("〜", 1)
        return parts[0].strip(), parts[1].strip()
    return v, ""

class _Conn:
    """psycopg2接続をSQLite風インターフェースでラップするクラス"""
    def __init__(self):
        self._conn = psycopg2.connect(DATABASE_URL)

    def execute(self, sql, params=()):
        # SQLiteの ? プレースホルダーをPostgreSQLの %s に変換
        sql = sql.replace("?", "%s")
        cur = self._conn.cursor(cursor_factory=psycopg2.extras.RealDictCursor)
        cur.execute(sql, params)
        return cur

    def __enter__(self):
        return self

    def __exit__(self, exc_type, *_):
        if exc_type:
            self._conn.rollback()
        else:
            self._conn.commit()
        self._conn.close()

def get_db():
    return _Conn()

def init_db():
    with get_db() as conn:
        conn.execute("""
            CREATE TABLE IF NOT EXISTS orders (
                id SERIAL PRIMARY KEY,
                order_no TEXT NOT NULL,
                customer TEXT NOT NULL,
                product TEXT NOT NULL,
                part_no TEXT NOT NULL DEFAULT '',
                material TEXT NOT NULL DEFAULT '',
                quantity INTEGER NOT NULL,
                mekki_type TEXT NOT NULL,
                due_date TEXT NOT NULL,
                note TEXT,
                assigned_to TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS customers (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS products (
                id SERIAL PRIMARY KEY,
                name TEXT NOT NULL UNIQUE,
                part_no TEXT NOT NULL DEFAULT '',
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS users (
                id SERIAL PRIMARY KEY,
                username TEXT NOT NULL UNIQUE,
                password_hash TEXT NOT NULL,
                created_at TEXT NOT NULL
            )
        """)
        conn.execute("""
            CREATE TABLE IF NOT EXISTS order_items (
                id SERIAL PRIMARY KEY,
                order_id INTEGER NOT NULL,
                part_no TEXT NOT NULL DEFAULT '',
                product TEXT NOT NULL,
                material TEXT NOT NULL DEFAULT '',
                quantity INTEGER NOT NULL,
                unit_price TEXT NOT NULL DEFAULT '',
                note TEXT NOT NULL DEFAULT ''
            )
        """)
        # 新カラムのマイグレーション（既存テーブルへの追加）
        for col, definition in [
            ("mekki_thickness",  "TEXT NOT NULL DEFAULT ''"),
            ("thickness_data",   "TEXT NOT NULL DEFAULT '不要'"),
            ("unit_price",       "TEXT NOT NULL DEFAULT ''"),
            ("mekki_line",       "TEXT NOT NULL DEFAULT ''"),
            ("process_note",     "TEXT NOT NULL DEFAULT ''"),
            ("shipping_method",  "TEXT NOT NULL DEFAULT ''"),
        ]:
            exists = conn.execute(
                "SELECT 1 FROM information_schema.columns WHERE table_name='orders' AND column_name=?",
                (col,)
            ).fetchone()
            if not exists:
                conn.execute(f"ALTER TABLE orders ADD COLUMN {col} {definition}")

        # デフォルトユーザーが未登録なら作成
        if not conn.execute("SELECT id FROM users WHERE username='admin'").fetchone():
            conn.execute(
                "INSERT INTO users (username, password_hash, created_at) VALUES (?, ?, ?)",
                ("admin", generate_password_hash("admin1234", method="pbkdf2:sha256"), datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
            )

@app.before_request
def check_session_timeout():
    if session.get("user_id"):
        last_active = session.get("last_active")
        if last_active:
            elapsed = datetime.now() - datetime.fromisoformat(last_active)
            if elapsed > timedelta(minutes=SESSION_TIMEOUT):
                session.clear()
                flash(f"{SESSION_TIMEOUT}分間操作がなかったため自動ログアウトしました。", "info")
                return redirect(url_for("login"))
        session["last_active"] = datetime.now().isoformat()

def login_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not session.get("user_id"):
            return redirect(url_for("login"))
        return f(*args, **kwargs)
    return decorated

# ── 認証 ────────────────────────────────────────────────────

@app.route("/login", methods=["GET", "POST"])
def login():
    if session.get("user_id"):
        return redirect(url_for("index"))
    error = None
    if request.method == "POST":
        username = request.form.get("username", "").strip()
        password = request.form.get("password", "")
        with get_db() as conn:
            user = conn.execute(
                "SELECT * FROM users WHERE username=?", (username,)
            ).fetchone()
        if user and check_password_hash(user["password_hash"], password):
            session.permanent = True
            session["user_id"] = user["id"]
            session["username"] = user["username"]
            session["last_active"] = datetime.now().isoformat()
            return redirect(url_for("index"))
        error = "ユーザー名またはパスワードが正しくありません。"
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    return redirect(url_for("login"))

@app.route("/change-password", methods=["GET", "POST"])
@login_required
def change_password():
    error = None
    success = None
    if request.method == "POST":
        current = request.form.get("current_password", "")
        new_pw  = request.form.get("new_password", "")
        confirm = request.form.get("confirm_password", "")
        with get_db() as conn:
            user = conn.execute("SELECT * FROM users WHERE id=?", (session["user_id"],)).fetchone()
        if not check_password_hash(user["password_hash"], current):
            error = "現在のパスワードが正しくありません。"
        elif len(new_pw) < 6:
            error = "新しいパスワードは6文字以上で入力してください。"
        elif new_pw != confirm:
            error = "新しいパスワードと確認用パスワードが一致しません。"
        else:
            with get_db() as conn:
                conn.execute(
                    "UPDATE users SET password_hash=? WHERE id=?",
                    (generate_password_hash(new_pw, method="pbkdf2:sha256"), session["user_id"])
                )
            success = "パスワードを変更しました。"
    return render_template("change_password.html", error=error, success=success)

@app.route("/")
@login_required
def index():
    with get_db() as conn:
        orders = conn.execute(
            "SELECT * FROM orders ORDER BY created_at DESC"
        ).fetchall()
    return render_template("index.html", orders=orders)

@app.route("/new", methods=["GET", "POST"])
@login_required
def new_order():
    if request.method == "POST":
        now = datetime.now()
        order_no = f"ORD-{now.strftime('%Y%m%d%H%M%S')}"
        with get_db() as conn:
            conn.execute("""
                INSERT INTO orders (order_no, customer, product, part_no, material, quantity,
                    mekki_type, mekki_thickness, thickness_data, due_date,
                    unit_price, mekki_line, process_note, shipping_method, note, assigned_to, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                order_no,
                request.form["customer"],
                request.form["product"],
                request.form.get("part_no", ""),
                request.form.get("material", ""),
                request.form["quantity"],
                request.form["mekki_type"],
                build_thickness(request.form.get("thickness_from", ""), request.form.get("thickness_to", "")),
                request.form.get("thickness_data", "不要"),
                request.form["due_date"],
                request.form.get("unit_price", ""),
                request.form.get("mekki_line", ""),
                request.form.get("process_note", ""),
                request.form.get("shipping_method", ""),
                request.form.get("note", ""),
                request.form.get("assigned_to", ""),
                now.strftime("%Y-%m-%d %H:%M:%S"),
            ))
        return redirect(url_for("index"))
    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
        products = conn.execute("SELECT name FROM products ORDER BY name").fetchall()
    return render_template("new.html", mekki_types=MEKKI_TYPES, mekki_lines=MEKKI_LINES,
                           customers=customers, products=products)

@app.route("/edit/<int:order_id>", methods=["GET", "POST"])
@login_required
def edit_order(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return redirect(url_for("index"))
    if request.method == "POST":
        with get_db() as conn:
            conn.execute("""
                UPDATE orders SET customer=?, product=?, part_no=?, material=?, quantity=?,
                mekki_type=?, mekki_thickness=?, thickness_data=?, due_date=?,
                unit_price=?, mekki_line=?, process_note=?, shipping_method=?, note=?, assigned_to=? WHERE id=?
            """, (
                request.form["customer"],
                request.form["product"],
                request.form.get("part_no", ""),
                request.form.get("material", ""),
                request.form["quantity"],
                request.form["mekki_type"],
                build_thickness(request.form.get("thickness_from", ""), request.form.get("thickness_to", "")),
                request.form.get("thickness_data", "不要"),
                request.form["due_date"],
                request.form.get("unit_price", ""),
                request.form.get("mekki_line", ""),
                request.form.get("process_note", ""),
                request.form.get("shipping_method", ""),
                request.form.get("note", ""),
                request.form.get("assigned_to", ""),
                order_id,
            ))
        return redirect(url_for("detail", order_id=order_id))
    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
        products = conn.execute("SELECT name FROM products ORDER BY name").fetchall()
    t_from, t_to = parse_thickness(order["mekki_thickness"])
    return render_template("edit.html", order=order, mekki_types=MEKKI_TYPES, mekki_lines=MEKKI_LINES,
                           customers=customers, products=products,
                           thickness_from=t_from, thickness_to=t_to)

@app.route("/detail/<int:order_id>")
@login_required
def detail(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return redirect(url_for("index"))
    return render_template("detail.html", order=order)

@app.route("/print/<int:order_id>")
@login_required
def print_order(order_id):
    with get_db() as conn:
        order = conn.execute("SELECT * FROM orders WHERE id=?", (order_id,)).fetchone()
    if not order:
        return redirect(url_for("index"))
    return render_template("print.html", order=order)

@app.route("/delete/<int:order_id>", methods=["POST"])
@login_required
def delete(order_id):
    with get_db() as conn:
        conn.execute("DELETE FROM orders WHERE id=?", (order_id,))
    return redirect(url_for("index"))

# ── 顧客マスタ ──────────────────────────────────────────────

@app.route("/customers", methods=["GET", "POST"])
@login_required
def customers():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.endswith(".xlsx"):
            flash("xlsx ファイルを選択してください。", "error")
            return redirect(url_for("customers"))
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            imported = 0
            skipped = 0
            with get_db() as conn:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = str(row[0]).strip() if row[0] is not None else ""
                    if not name or name == "None":
                        continue
                    existing = conn.execute(
                        "SELECT id FROM customers WHERE name=?", (name,)
                    ).fetchone()
                    if existing:
                        skipped += 1
                    else:
                        conn.execute(
                            "INSERT INTO customers (name, created_at) VALUES (?, ?)",
                            (name, now)
                        )
                        imported += 1
            msg = f"{imported} 件を登録しました。"
            if skipped:
                msg += f"（{skipped} 件は既存のためスキップ）"
            flash(msg, "success")
        except Exception as e:
            flash(f"読み込みエラー: {e}", "error")
        return redirect(url_for("customers"))

    with get_db() as conn:
        customers_list = conn.execute(
            "SELECT * FROM customers ORDER BY name"
        ).fetchall()
    return render_template("customers.html", customers=customers_list)

@app.route("/customers/edit/<int:customer_id>", methods=["POST"])
@login_required
def edit_customer(customer_id):
    name = request.form.get("name", "").strip()
    if name:
        try:
            with get_db() as conn:
                conn.execute("UPDATE customers SET name=? WHERE id=?", (name, customer_id))
            flash(f"顧客名を「{name}」に更新しました。", "success")
        except Exception:
            flash("同じ顧客名が既に登録されています。", "error")
    return redirect(url_for("customers"))

@app.route("/customers/delete/<int:customer_id>", methods=["POST"])
@login_required
def delete_customer(customer_id):
    with get_db() as conn:
        conn.execute("DELETE FROM customers WHERE id=?", (customer_id,))
    return redirect(url_for("customers"))

@app.route("/customers/delete-all", methods=["POST"])
@login_required
def delete_all_customers():
    with get_db() as conn:
        conn.execute("DELETE FROM customers")
    flash("顧客マスタを全件削除しました。", "success")
    return redirect(url_for("customers"))

# ── 品名マスタ ──────────────────────────────────────────────

@app.route("/products", methods=["GET", "POST"])
@login_required
def products():
    if request.method == "POST":
        file = request.files.get("file")
        if not file or not file.filename.endswith(".xlsx"):
            flash("xlsx ファイルを選択してください。", "error")
            return redirect(url_for("products"))
        try:
            wb = openpyxl.load_workbook(file)
            ws = wb.active
            now = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            imported = 0
            skipped = 0
            with get_db() as conn:
                for row in ws.iter_rows(min_row=2, values_only=True):
                    name = str(row[0]).strip() if row[0] is not None else ""
                    if not name or name == "None":
                        continue
                    existing = conn.execute(
                        "SELECT id FROM products WHERE name=?", (name,)
                    ).fetchone()
                    if existing:
                        skipped += 1
                    else:
                        conn.execute(
                            "INSERT INTO products (name, created_at) VALUES (?, ?)",
                            (name, now)
                        )
                        imported += 1
            msg = f"{imported} 件を登録しました。"
            if skipped:
                msg += f"（{skipped} 件は既存のためスキップ）"
            flash(msg, "success")
        except Exception as e:
            flash(f"読み込みエラー: {e}", "error")
        return redirect(url_for("products"))

    with get_db() as conn:
        products_list = conn.execute(
            "SELECT * FROM products ORDER BY name"
        ).fetchall()
    return render_template("products.html", products=products_list)

@app.route("/products/add", methods=["POST"])
@login_required
def add_product():
    name = request.form.get("name", "").strip()
    if name:
        try:
            part_no = request.form.get("part_no", "").strip()
            with get_db() as conn:
                conn.execute(
                    "INSERT INTO products (name, part_no, created_at) VALUES (?, ?, ?)",
                    (name, part_no, datetime.now().strftime("%Y-%m-%d %H:%M:%S"))
                )
            flash(f"「{name}」を登録しました。", "success")
        except Exception:
            flash("同じ品名が既に登録されています。", "error")
    return redirect(url_for("products"))

@app.route("/products/edit/<int:product_id>", methods=["POST"])
@login_required
def edit_product(product_id):
    name = request.form.get("name", "").strip()
    if name:
        try:
            new_part_no = request.form.get("part_no", "").strip()
            with get_db() as conn:
                conn.execute("UPDATE products SET name=?, part_no=? WHERE id=?", (name, new_part_no, product_id))
            flash(f"品名を「{name}」に更新しました。", "success")
        except Exception:
            flash("同じ品名が既に登録されています。", "error")
    return redirect(url_for("products"))

@app.route("/products/delete/<int:product_id>", methods=["POST"])
@login_required
def delete_product(product_id):
    with get_db() as conn:
        conn.execute("DELETE FROM products WHERE id=?", (product_id,))
    return redirect(url_for("products"))

@app.route("/products/delete-all", methods=["POST"])
@login_required
def delete_all_products():
    with get_db() as conn:
        conn.execute("DELETE FROM products")
    flash("品名マスタを全件削除しました。", "success")
    return redirect(url_for("products"))

# ── Excelエクスポート ────────────────────────────────────────

@app.route("/export/excel")
@login_required
def export_excel():
    with get_db() as conn:
        orders = conn.execute(
            "SELECT * FROM orders ORDER BY created_at DESC"
        ).fetchall()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "受注データ"

    headers = ["伝票番号", "顧客名", "品名", "品番", "材質", "数量", "めっき種類", "めっき膜厚", "膜厚データ", "納期", "単価", "めっきライン", "工程", "出荷方法", "備考", "登録日時"]
    ws.append(headers)

    # ヘッダー行のスタイル
    header_fill = PatternFill(fill_type="solid", fgColor="1A3A5C")
    header_font = Font(bold=True, color="FFFFFF", size=11)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center")

    # データ行
    for order in orders:
        ws.append([
            order["order_no"],
            order["customer"],
            order["product"],
            order["part_no"] or "",
            order["material"] or "",
            order["quantity"],
            order["mekki_type"],
            order["mekki_thickness"] or "",
            order["thickness_data"] or "",
            order["due_date"],
            order["unit_price"] or "",
            order["mekki_line"] or "",
            order["process_note"] or "",
            order["shipping_method"] or "",
            order["note"] or "",
            order["created_at"],
        ])

    # 列幅の自動調整
    col_widths = [22, 20, 20, 16, 16, 8, 16, 14, 10, 14, 10, 16, 24, 12, 30, 20]
    for i, width in enumerate(col_widths, start=1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = width

    buf = BytesIO()
    wb.save(buf)
    buf.seek(0)

    filename = f"orders_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(
        buf,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )

# ─── 複数品目受注 ───────────────────────────────────────

@app.route("/edit_multi/<int:order_id>", methods=["GET", "POST"])
@login_required
def edit_order_multi(order_id):
    with get_db() as conn:
        order = conn.execute(
            "SELECT * FROM orders WHERE id=?", (order_id,)
        ).fetchone()
        items = conn.execute(
            "SELECT * FROM order_items WHERE order_id=? ORDER BY id",
            (order_id,)
        ).fetchall()
    if not order:
        return redirect(url_for("index"))

    if request.method == "POST":
        with get_db() as conn:
            # ヘッダー更新
            conn.execute("""
                UPDATE orders SET
                    customer=?, mekki_type=?, mekki_thickness=?, thickness_data=?,
                    material=?, due_date=?, mekki_line=?, process_note=?,
                    shipping_method=?, note=?, assigned_to=?
                WHERE id=?
            """, (
                request.form["customer"],
                request.form["mekki_type"],
                build_thickness(request.form.get("thickness_from", ""), request.form.get("thickness_to", "")),
                request.form.get("thickness_data", "不要"),
                request.form.get("material", ""),
                request.form["due_date"],
                request.form.get("mekki_line", ""),
                request.form.get("process_note", ""),
                request.form.get("shipping_method", ""),
                request.form.get("note", ""),
                request.form.get("assigned_to", ""),
                order_id,
            ))
            # 既存の明細を全削除して再登録
            conn.execute("DELETE FROM order_items WHERE order_id=?", (order_id,))
            products = request.form.getlist("product[]")
            part_nos = request.form.getlist("part_no[]")
            materials = request.form.getlist("material[]")
            quantities = request.form.getlist("quantity[]")
            unit_prices = request.form.getlist("unit_price[]")
            notes = request.form.getlist("note[]")
            for i, product in enumerate(products):
                if not product.strip():
                    continue
                conn.execute("""
                    INSERT INTO order_items (
                        order_id, part_no, product, material, quantity, unit_price, note
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    order_id,
                    part_nos[i] if i < len(part_nos) else "",
                    product,
                    materials[i] if i < len(materials) else "",
                    int(quantities[i]) if i < len(quantities) and quantities[i] else 0,
                    unit_prices[i] if i < len(unit_prices) else "",
                    notes[i] if i < len(notes) else "",
                ))
        return redirect(url_for("detail_multi", order_id=order_id))

    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
        products = conn.execute("SELECT name, part_no FROM products ORDER BY name").fetchall()
    t_from, t_to = parse_thickness(order["mekki_thickness"])
    return render_template("edit_multi.html",
                           order=order,
                           items=items,
                           mekki_types=MEKKI_TYPES,
                           mekki_lines=MEKKI_LINES,
                           customers=customers,
                           products=products,
                           thickness_from=t_from,
                           thickness_to=t_to)


@app.route("/new_multi", methods=["GET", "POST"])
@login_required
def new_order_multi():
    if request.method == "POST":
        now = datetime.now()
        order_no = f"ORD-{now.strftime('%Y%m%d%H%M%S')}-M"
        with get_db() as conn:
            # ヘッダー情報をordersテーブルに登録（product="複数品目"として保存）
            conn.execute("""
                INSERT INTO orders (
                    order_no, customer, product, part_no, material, quantity,
                    mekki_type, mekki_thickness, thickness_data, due_date,
                    unit_price, mekki_line, process_note, shipping_method, note, assigned_to, created_at
                ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
            """, (
                order_no,
                request.form["customer"],
                "複数品目",
                "",
                request.form.get("material", ""),
                0,
                request.form["mekki_type"],
                build_thickness(request.form.get("thickness_from", ""), request.form.get("thickness_to", "")),
                request.form.get("thickness_data", "不要"),
                request.form["due_date"],
                "",
                request.form.get("mekki_line", ""),
                request.form.get("process_note", ""),
                request.form.get("shipping_method", ""),
                request.form.get("note", ""),
                request.form.get("assigned_to", ""),
                now.strftime("%Y-%m-%d %H:%M:%S"),
            ))
            order = conn.execute(
                "SELECT * FROM orders WHERE order_no=?", (order_no,)
            ).fetchone()
            order_id = order["id"]

            # 明細行をorder_itemsテーブルに登録
            products = request.form.getlist("product[]")
            part_nos = request.form.getlist("part_no[]")
            materials = request.form.getlist("material[]")
            quantities = request.form.getlist("quantity[]")
            unit_prices = request.form.getlist("unit_price[]")
            notes = request.form.getlist("note[]")

            for i, product in enumerate(products):
                if not product.strip():
                    continue
                conn.execute("""
                    INSERT INTO order_items (
                        order_id, part_no, product, material, quantity, unit_price, note
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                """, (
                    order_id,
                    part_nos[i] if i < len(part_nos) else "",
                    product,
                    materials[i] if i < len(materials) else "",
                    int(quantities[i]) if i < len(quantities) and quantities[i] else 0,
                    unit_prices[i] if i < len(unit_prices) else "",
                    notes[i] if i < len(notes) else "",
                ))
        return redirect(url_for("index"))

    with get_db() as conn:
        customers = conn.execute("SELECT name FROM customers ORDER BY name").fetchall()
        products = conn.execute("SELECT name, part_no FROM products ORDER BY name").fetchall()
    return render_template("new_multi.html",
                           mekki_types=MEKKI_TYPES,
                           mekki_lines=MEKKI_LINES,
                           customers=customers,
                           products=products)


@app.route("/detail_multi/<int:order_id>")
@login_required
def detail_multi(order_id):
    with get_db() as conn:
        order = conn.execute(
            "SELECT * FROM orders WHERE id=?", (order_id,)
        ).fetchone()
        items = conn.execute(
            "SELECT * FROM order_items WHERE order_id=? ORDER BY id",
            (order_id,)
        ).fetchall()
    if not order:
        return redirect(url_for("index"))
    return render_template("detail_multi.html", order=order, items=items)


@app.route("/print_multi/<int:order_id>")
@login_required
def print_order_multi(order_id):
    with get_db() as conn:
        order = conn.execute(
            "SELECT * FROM orders WHERE id=?", (order_id,)
        ).fetchone()
        items = conn.execute(
            "SELECT * FROM order_items WHERE order_id=? ORDER BY id",
            (order_id,)
        ).fetchall()
    if not order:
        return redirect(url_for("index"))
    return render_template("print_multi.html", order=order, items=items)


init_db()

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5001, debug=True)
